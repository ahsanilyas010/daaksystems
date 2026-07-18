#!/usr/bin/env python3
"""Clean CUSTOMER_PARCELS_DETAIL.xlsx into normalized CSVs for Postgres import.

Reads the 8 raw carrier/agent sheets, applies the field-level normalization
rules described in plan.md section 7 Phase 0, and writes:
  - customers.csv, cities.csv, shipments.csv, shipment_events.csv
  - data_quality_report.csv (every row-level issue found, for manual review)
  - summary.txt (counts, so a human can sanity-check the run at a glance)

Usage:
    python3 scripts/migrate.py [--input PATH] [--outdir DIR]
"""
from __future__ import annotations

import argparse
import csv
import difflib
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = REPO_ROOT / "data" / "raw" / "CUSTOMER_PARCELS_DETAIL.xlsx"
DEFAULT_OUTDIR = REPO_ROOT / "data" / "processed"

EXCEL_EPOCH = date(1899, 12, 30)  # Excel's day-0, accounting for the 1900 leap-year bug


# --------------------------------------------------------------------------
# Sheet configuration — maps each sheet's raw columns to canonical fields.
# Column indices are 1-based, matching openpyxl.
# --------------------------------------------------------------------------
@dataclass(frozen=True)
class SheetConfig:
    sheet_name: str
    header_row: int
    carrier_code: str
    col_sender: int | None       # who booked/owns the parcel (a Daak customer)
    col_consignee: int | None    # who receives the parcel
    col_tracking: int | None
    col_city: int | None
    col_weight: int | None
    col_cod_amount: int | None   # "PARCEL AMOUNT" / "AMOUNT" / "AMT"
    col_dc: int | None           # delivery charge billed to customer
    col_status: int | None
    col_rider: int | None
    col_carrier_cost: int | None = None
    fixed_sender: str | None = None      # sheet has no real sender column
    status_unreliable: bool = False      # STATUS column doesn't hold delivery status


SHEETS = [
    SheetConfig(
        sheet_name="ENTRY SHEET DAAK CALL COURIER", header_row=4, carrier_code="call_courier",
        col_sender=3, col_consignee=4, col_tracking=5, col_city=6, col_weight=7,
        col_cod_amount=8, col_dc=9, col_status=11, col_rider=13,
        fixed_sender="DAAK",
    ),
    SheetConfig(
        sheet_name="ENTRY SHEET DAAK BLUE X", header_row=4, carrier_code="bluex",
        col_sender=3, col_consignee=4, col_tracking=5, col_city=6, col_weight=7,
        col_cod_amount=8, col_dc=9, col_status=11, col_rider=13, col_carrier_cost=15,
    ),
    SheetConfig(
        sheet_name="ARIF BHAI", header_row=3, carrier_code="arif_bhai",
        col_sender=3, col_consignee=4, col_tracking=None, col_city=None, col_weight=None,
        col_cod_amount=5, col_dc=6, col_status=8, col_rider=None,
    ),
    SheetConfig(
        sheet_name="ROCKET", header_row=4, carrier_code="rocket",
        col_sender=3, col_consignee=4, col_tracking=5, col_city=6, col_weight=7,
        col_cod_amount=8, col_dc=9, col_status=11, col_rider=None,
    ),
    SheetConfig(
        sheet_name="COURIER NEXT", header_row=1, carrier_code="courier_next",
        col_sender=3, col_consignee=4, col_tracking=5, col_city=6, col_weight=7,
        col_cod_amount=8, col_dc=9, col_status=11, col_rider=None,
    ),
    SheetConfig(
        sheet_name="LEOPARD", header_row=4, carrier_code="leopard",
        col_sender=5, col_consignee=4, col_tracking=3, col_city=10, col_weight=6,
        col_cod_amount=8, col_dc=7, col_status=11, col_rider=None,
        status_unreliable=True,
    ),
    SheetConfig(
        sheet_name="DO DELIVER", header_row=4, carrier_code="dodeliver",
        col_sender=3, col_consignee=4, col_tracking=5, col_city=6, col_weight=7,
        col_cod_amount=8, col_dc=9, col_status=11, col_rider=None,
    ),
    SheetConfig(
        sheet_name="HASSAN", header_row=3, carrier_code="hassan",
        col_sender=3, col_consignee=4, col_tracking=None, col_city=None, col_weight=None,
        col_cod_amount=5, col_dc=6, col_status=8, col_rider=None,
    ),
]

# --------------------------------------------------------------------------
# Status normalization — locked target enum from plan.md section 3.
# --------------------------------------------------------------------------
STATUS_MAP = {
    "DELIVER": "DELIVERED",
    "DELIVERED": "DELIVERED",
    "D": "DELIVERED",
    "RETURN": "RETURNED",
    "RETURNED": "RETURNED",
    "LOST": "LOST",
    "PICK": "PICKED",
    "PICKED": "PICKED",
    "REVERSE": "RETURN_INITIATED",
    "MISTAKE": "CANCELLED",
    "CANCELLED": "CANCELLED",
    "CANCEL": "CANCELLED",
}
# Raw values seen in the sheets that are not a delivery status at all
# (payment confirmations, ambiguous ops notes) -> left unmapped, flagged.
KNOWN_AMBIGUOUS_STATUS = {"MISSED", "ROUTE ISSUE", "MISS", "REF", "MS", "PAID"}

RIDER_MAP = {"WAL": "WAL", "SAL": "SAL", "ASFAR": "ASFAR", "SHB": "SHB"}
# Free-text noise observed in the RIDER column (see plan.md section 1).
KNOWN_RIDER_NOISE = {"CLEAR", "OK", "P", ""}

CITY_ALIASES = {
    "KHI": ("Karachi", "KHI"), "KARACHI": ("Karachi", "KHI"),
    "LHR": ("Lahore", "LHR"), "LAHORE": ("Lahore", "LHR"),
    "ISB": ("Islamabad", "ISB"), "ISLAMABAD": ("Islamabad", "ISB"),
    "RWP": ("Rawalpindi", "RWP"), "RAWALPINDI": ("Rawalpindi", "RWP"),
    "FSB": ("Faisalabad", "FSB"), "FAISALABAD": ("Faisalabad", "FSB"),
    "PESH": ("Peshawar", "PESH"), "PESHAWAR": ("Peshawar", "PESH"),
    "HYD": ("Hyderabad", "HYD"), "HYDERABAD": ("Hyderabad", "HYD"), "HDD": ("Hyderabad", "HYD"),
    "MULTAN": ("Multan", "MUL"),
    "GUJ": ("Gujranwala", "GUJ"), "GUJRANWALA": ("Gujranwala", "GUJ"),
    "GRT": ("Gujrat", "GRT"), "GUJRAT": ("Gujrat", "GRT"),
    "SKT": ("Sialkot", "SKT"), "SIALKOT": ("Sialkot", "SKT"),
    "QUETTA": ("Quetta", "QTA"), "QTA": ("Quetta", "QTA"),
    "SARGODHA": ("Sargodha", "SGD"), "SRG": ("Sargodha", "SGD"),
    "MUZZ": ("Muzaffargarh", "MUZ"), "MUZAFFARGARH": ("Muzaffargarh", "MUZ"),
    "DERA ISMAIL": ("Dera Ismail Khan", "DIK"), "DIK": ("Dera Ismail Khan", "DIK"),
    "KHANEWAL": ("Khanewal", "KHN"),
    "MIRPUR": ("Mirpur", "MPR"),
    "BAGH": ("Bagh", "BAG"),
    "HASILPUR": ("Hasilpur", "HSP"),
}
# Sheet noise seen literally in the CITY column that is not a city at all.
KNOWN_NON_CITY_TOKENS = {"-", "TCS", "", "YOU", "SH", "WAL", "SOH"}

# Sender-name variants collapsed deterministically (honorific/whitespace/version
# suffixes only — see normalize_customer_key). Anything not covered here that
# merely *looks* similar to an existing customer is flagged, not auto-merged.
CUSTOMER_ALIASES = {
    "AK": "AK COLLECTION",
    "AK HYD": "AK COLLECTION",
    "DOWN": "DOWN TOWN",
}

TRACKING_MIN_LEN = 5


def excel_serial_to_date(serial: float) -> date:
    return EXCEL_EPOCH + timedelta(days=serial)


def normalize_date(raw) -> tuple[date | None, str | None]:
    if raw is None:
        return None, "missing date"
    if isinstance(raw, datetime):
        return raw.date(), None
    if isinstance(raw, date):
        return raw, None
    if isinstance(raw, (int, float)):
        try:
            return excel_serial_to_date(float(raw)), None
        except (OverflowError, ValueError):
            return None, f"unparseable excel serial date: {raw!r}"
    if isinstance(raw, str):
        s = raw.strip()
        if not s:
            return None, "missing date"
        m = re.match(r"^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})$", s)
        if m:
            d, mo, y = (int(g) for g in m.groups())
            if y < 100:
                y += 2000
            try:
                return date(y, mo, d), None
            except ValueError:
                return None, f"unparseable date string: {raw!r}"
        return None, f"unparseable date string: {raw!r}"
    return None, f"unexpected date type: {raw!r}"


def normalize_status(raw, status_unreliable: bool) -> tuple[str | None, str | None]:
    if status_unreliable:
        return None, f"status column not trustworthy for this sheet (raw={raw!r})"
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None, "missing status"
    key = str(raw).strip().upper()
    if key in STATUS_MAP:
        return STATUS_MAP[key], None
    if key in KNOWN_AMBIGUOUS_STATUS:
        return None, f"ambiguous/non-delivery status text: {raw!r}"
    return None, f"unrecognized status text: {raw!r}"


def normalize_rider(raw) -> tuple[str | None, str | None]:
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None, None  # no rider recorded is common and not itself an error
    key = str(raw).strip().upper()
    if key in RIDER_MAP:
        return RIDER_MAP[key], None
    if key in KNOWN_RIDER_NOISE:
        return None, f"rider column contains non-rider text: {raw!r}"
    return None, f"unrecognized rider code: {raw!r}"


def normalize_tracking(raw) -> tuple[str | None, str | None]:
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None, "missing tracking number"
    s = str(raw).strip()
    if isinstance(raw, float) and raw.is_integer():
        s = str(int(raw))
    key = s.upper()
    first_token = key.split()[0] if key.split() else key
    has_digit_run = re.search(r"\d{4,}", key) is not None
    if key in CITY_ALIASES or key in KNOWN_NON_CITY_TOKENS or len(s) < TRACKING_MIN_LEN:
        return None, f"invalid tracking number (looks like noise): {raw!r}"
    if first_token in CITY_ALIASES and not has_digit_run:
        return None, f"invalid tracking number (looks like city+rider text, not a tracking no): {raw!r}"
    return s, None


class CityRegistry:
    def __init__(self):
        self._by_key: dict[str, dict] = {}  # code -> city record
        self._next_id = 1
        self._resolved_raw_cache: dict[str, tuple[dict | None, str | None]] = {}
        # raw value (upper) -> {"count": n, "resolved_name": str, "note": str}, for a rollup report
        self.review: dict[str, dict] = {}

    def resolve(self, raw) -> tuple[dict | None, str | None]:
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            return None, "missing city"
        s = str(raw).strip()
        key = s.upper()

        if key in self._resolved_raw_cache:
            rec, note = self._resolved_raw_cache[key]
            if note:
                self._record_review(key, rec, note)
            return rec, None  # already surfaced in the review rollup on first sighting

        rec, note = self._resolve_uncached(s, key)
        self._resolved_raw_cache[key] = (rec, note)
        if note:
            self._record_review(key, rec, note)
        return rec, None

    def _record_review(self, raw_key: str, rec: dict | None, note: str):
        entry = self.review.setdefault(raw_key, {
            "raw_value": raw_key, "resolved_name": rec["name"] if rec else "(unresolved)",
            "note": note, "count": 0,
        })
        entry["count"] += 1

    def _resolve_uncached(self, s: str, key: str) -> tuple[dict | None, str | None]:
        if key in CITY_ALIASES:
            name, code = CITY_ALIASES[key]
            return self._get_or_create(code, name), None

        # Strip a trailing noise token, e.g. "KHI SH" -> "KHI" (+ "SH" flagged).
        parts = key.split()
        if len(parts) > 1 and parts[0] in CITY_ALIASES:
            name, code = CITY_ALIASES[parts[0]]
            note = f"city value had trailing text stripped to resolve to {name}"
            return self._get_or_create(code, name), note

        if key in KNOWN_NON_CITY_TOKENS:
            return None, "unrecognized city value"

        # Fuzzy match against canonical names already resolved so far.
        canon_names = [c["name"] for c in self._by_key.values()]
        close = difflib.get_close_matches(s.title(), canon_names, n=1, cutoff=0.86)
        if close:
            match = next(c for c in self._by_key.values() if c["name"] == close[0])
            note = f"fuzzy-matched to existing city {match['name']} — verify this is correct"
            return match, note

        # Unknown city: register it as its own row but flag for human review.
        code = re.sub(r"[^A-Z0-9]", "", key)[:6] or "UNK"
        rec = self._get_or_create(code, s.title())
        return rec, "new/uncertain city value, added as its own row — verify/merge"

    def _get_or_create(self, code: str, name: str) -> dict:
        if code not in self._by_key:
            self._by_key[code] = {"id": self._next_id, "name": name, "code": code}
            self._next_id += 1
        return self._by_key[code]

    def rows(self):
        return sorted(self._by_key.values(), key=lambda c: c["id"])


HONORIFIC_RE = re.compile(r"^(MR|MRS|MS|SYED|SYEDA|ST)\.?\s+")
TRAILING_VERSION_RE = re.compile(r"\s+\d+(\.\d+)?$")


def normalize_customer_key(raw: str) -> str:
    s = re.sub(r"\s+", " ", raw.strip().upper())
    s = HONORIFIC_RE.sub("", s)
    s = TRAILING_VERSION_RE.sub("", s)
    return s.strip()


class CustomerRegistry:
    def __init__(self):
        self._by_key: dict[str, dict] = {}
        self._next_id = 1
        self.possible_duplicates: list[tuple[str, str]] = []

    def resolve(self, raw: str) -> dict:
        display = re.sub(r"\s+", " ", raw.strip())
        alias_target = CUSTOMER_ALIASES.get(display.upper())
        key = normalize_customer_key(alias_target or display)

        if key not in self._by_key:
            close = difflib.get_close_matches(key, self._by_key.keys(), n=1, cutoff=0.90)
            if close:
                self.possible_duplicates.append((display, self._by_key[close[0]]["name"]))
            self._by_key[key] = {
                "id": self._next_id,
                "name": alias_target or display,
                "aliases": set(),
            }
            self._next_id += 1
        rec = self._by_key[key]
        rec["aliases"].add(display)
        return rec

    def rows(self):
        return sorted(self._by_key.values(), key=lambda c: c["id"])


def cell(ws, row: int, col: int | None):
    if col is None:
        return None
    return ws.cell(row=row, column=col).value


def iter_data_rows(ws, cfg: SheetConfig):
    for r in range(cfg.header_row + 1, ws.max_row + 1):
        probe_cols = [c for c in (cfg.col_sender, cfg.col_consignee, cfg.col_tracking) if c]
        if not any(cell(ws, r, c) not in (None, "") for c in probe_cols):
            continue  # fully-blank spacer row
        yield r


def run(input_path: Path, outdir: Path):
    outdir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(input_path, data_only=True)

    cities = CityRegistry()
    customers = CustomerRegistry()
    shipments: list[dict] = []
    events: list[dict] = []
    quality_report: list[dict] = []
    seen_tracking: dict[tuple[str, str], tuple[str, int]] = {}  # (carrier, tracking_no) -> first (sheet, row)
    skipped_rows = 0

    def flag(sheet, row, field, issue, raw_value):
        quality_report.append({
            "sheet": sheet, "row": row, "field": field,
            "issue": issue, "raw_value": "" if raw_value is None else str(raw_value),
        })

    shipment_id = 1
    event_id = 1

    for cfg in SHEETS:
        if cfg.sheet_name not in wb.sheetnames:
            raise ValueError(f"expected sheet not found in workbook: {cfg.sheet_name}")
        ws = wb[cfg.sheet_name]

        for r in iter_data_rows(ws, cfg):
            sender_raw = cfg.fixed_sender or cell(ws, r, cfg.col_sender)
            if not sender_raw or not str(sender_raw).strip():
                flag(cfg.sheet_name, r, "sender", "missing sender/customer name", sender_raw)
                skipped_rows += 1
                continue
            if cfg.fixed_sender:
                flag(cfg.sheet_name, r, "sender", "sheet has no real sender column; defaulted to DAAK", sender_raw)
            customer = customers.resolve(str(sender_raw))

            consignee_raw = cell(ws, r, cfg.col_consignee)
            if not consignee_raw or not str(consignee_raw).strip():
                flag(cfg.sheet_name, r, "consignee", "missing consignee name", consignee_raw)

            booked_at, date_issue = normalize_date(cell(ws, r, 2))
            if date_issue:
                flag(cfg.sheet_name, r, "date", date_issue, cell(ws, r, 2))

            tracking_no, tracking_issue = None, None
            if cfg.col_tracking:
                tracking_raw = cell(ws, r, cfg.col_tracking)
                tracking_no, tracking_issue = normalize_tracking(tracking_raw)
                if tracking_issue:
                    flag(cfg.sheet_name, r, "tracking_no", tracking_issue, tracking_raw)
                elif tracking_no:
                    dupe_key = (cfg.carrier_code, tracking_no)
                    if dupe_key in seen_tracking:
                        first_sheet, first_row = seen_tracking[dupe_key]
                        tracking_issue = (
                            f"duplicate tracking number for carrier {cfg.carrier_code!r}, "
                            f"also used by {first_sheet} row {first_row}: {tracking_raw!r}"
                        )
                        flag(cfg.sheet_name, r, "tracking_no", tracking_issue, tracking_raw)
                        tracking_no = None
                    else:
                        seen_tracking[dupe_key] = (cfg.sheet_name, r)

            city = None
            if cfg.col_city:
                city_raw = cell(ws, r, cfg.col_city)
                city, _ = cities.resolve(city_raw)  # per-value issues are rolled up in city_review.csv

            status_raw = cell(ws, r, cfg.col_status)
            status, status_issue = normalize_status(status_raw, cfg.status_unreliable)
            if status_issue:
                flag(cfg.sheet_name, r, "status", status_issue, status_raw)

            rider_raw = cell(ws, r, cfg.col_rider)
            rider_code, rider_issue = normalize_rider(rider_raw)
            if rider_issue:
                flag(cfg.sheet_name, r, "rider", rider_issue, rider_raw)

            weight = cell(ws, r, cfg.col_weight)
            weight_kg = float(weight) if isinstance(weight, (int, float)) else None
            if cfg.col_weight and weight_kg is None and weight not in (None, ""):
                flag(cfg.sheet_name, r, "weight", f"non-numeric weight: {weight!r}", weight)

            cod_amount = cell(ws, r, cfg.col_cod_amount)
            cod_amount = float(cod_amount) if isinstance(cod_amount, (int, float)) else None
            dc_amount = cell(ws, r, cfg.col_dc)
            dc_amount = float(dc_amount) if isinstance(dc_amount, (int, float)) else None

            carrier_cost = cell(ws, r, cfg.col_carrier_cost)
            carrier_cost = float(carrier_cost) if isinstance(carrier_cost, (int, float)) else None
            profit = (dc_amount - carrier_cost) if (dc_amount is not None and carrier_cost is not None) else None

            shipments.append({
                "id": shipment_id,
                "daak_tracking_no": f"DAAK-{booked_at.strftime('%y%m%d') if booked_at else '000000'}-{shipment_id:05d}",
                "carrier_code": cfg.carrier_code,
                "carrier_tracking_no": tracking_no,
                "customer_id": customer["id"],
                "consignee_name": str(consignee_raw).strip() if consignee_raw else None,
                "city_id": city["id"] if city else None,
                "weight_kg": weight_kg,
                "pieces": 1,
                "declared_value": None,
                "cod_amount": cod_amount,
                "dc_amount": dc_amount,
                "carrier_cost": carrier_cost,
                "profit": profit,
                "status": status,
                "booked_at": booked_at.isoformat() if booked_at else None,
                "rider_code": rider_code,
                "source_sheet": cfg.sheet_name,
                "source_row": r,
                "needs_review": bool(date_issue or tracking_issue or status_issue),
            })

            events.append({
                "id": event_id, "shipment_id": shipment_id, "status": "BOOKED",
                "source": "migration", "location": None, "note": "backfilled from historical sheet",
                "actor": None, "created_at": booked_at.isoformat() if booked_at else None,
            })
            event_id += 1
            if status:
                events.append({
                    "id": event_id, "shipment_id": shipment_id, "status": status,
                    "source": "migration", "location": None,
                    "note": "terminal status backfilled from historical sheet (exact transition dates not captured)",
                    "actor": None, "created_at": booked_at.isoformat() if booked_at else None,
                })
                event_id += 1

            shipment_id += 1

    write_csv(outdir / "customers.csv",
              ["id", "name", "phone", "email", "cnic", "address", "cod_payout_method",
               "credit_limit", "source_aliases"],
              [{
                  "id": c["id"], "name": c["name"], "phone": "", "email": "", "cnic": "",
                  "address": "", "cod_payout_method": "", "credit_limit": "",
                  "source_aliases": "; ".join(sorted(c["aliases"])),
              } for c in customers.rows()])

    write_csv(outdir / "cities.csv",
              ["id", "name", "code", "zone"],
              [{"id": c["id"], "name": c["name"], "code": c["code"], "zone": ""} for c in cities.rows()])

    write_csv(outdir / "shipments.csv",
              ["id", "daak_tracking_no", "carrier_code", "carrier_tracking_no", "customer_id",
               "consignee_name", "city_id", "weight_kg", "pieces", "declared_value", "cod_amount",
               "dc_amount", "carrier_cost", "profit", "status", "booked_at", "rider_code",
               "source_sheet", "source_row", "needs_review"],
              shipments)

    write_csv(outdir / "shipment_events.csv",
              ["id", "shipment_id", "status", "source", "location", "note", "actor", "created_at"],
              events)

    write_csv(outdir / "data_quality_report.csv",
              ["sheet", "row", "field", "issue", "raw_value"],
              quality_report)

    write_csv(outdir / "city_review.csv",
              ["raw_value", "resolved_name", "shipment_count", "note"],
              [{"raw_value": v["raw_value"], "resolved_name": v["resolved_name"],
                "shipment_count": v["count"], "note": v["note"]}
               for v in sorted(cities.review.values(), key=lambda v: -v["count"])])

    write_csv(outdir / "customer_review.csv",
              ["raw_value", "looks_similar_to", "note"],
              [{"raw_value": display, "looks_similar_to": matched_to,
                "note": "possible duplicate — verify before merging in customers.csv"}
               for display, matched_to in customers.possible_duplicates])

    write_summary(outdir / "summary.txt", cfgs=SHEETS, shipments=shipments, customers=customers,
                  cities=cities, quality_report=quality_report, skipped_rows=skipped_rows)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict]):
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_summary(path: Path, cfgs, shipments, customers, cities, quality_report, skipped_rows):
    by_sheet = {}
    for s in shipments:
        by_sheet.setdefault(s["source_sheet"], 0)
        by_sheet[s["source_sheet"]] += 1
    needs_review = sum(1 for s in shipments if s["needs_review"])
    issue_counts = {}
    for row in quality_report:
        issue_counts[row["field"]] = issue_counts.get(row["field"], 0) + 1

    lines = [
        "DAAK migration summary",
        "=======================",
        f"Shipments written: {len(shipments)}",
        f"Rows skipped (no sender/consignee/tracking at all): {skipped_rows}",
        f"Shipments flagged needs_review: {needs_review} ({needs_review / max(len(shipments),1):.1%})",
        f"Customers: {len(customers.rows())}",
        f"Cities: {len(cities.rows())}",
        f"data_quality_report.csv rows (per-shipment issues): {len(quality_report)}",
        f"city_review.csv rows (unique raw city values needing a human look): {len(cities.review)}",
        f"customer_review.csv rows (possible duplicate senders): {len(customers.possible_duplicates)}",
        "",
        "Shipments per sheet:",
    ]
    for sheet, count in by_sheet.items():
        lines.append(f"  {sheet}: {count}")
    lines.append("")
    lines.append("Sheets with no tracking-number column at all (agent ledgers, not a data-entry gap):")
    for cfg in cfgs:
        if cfg.col_tracking is None:
            lines.append(f"  {cfg.sheet_name}")
    lines.append("")
    lines.append("Sheets with no city column at all:")
    for cfg in cfgs:
        if cfg.col_city is None:
            lines.append(f"  {cfg.sheet_name}")
    lines.append("")
    lines.append("Per-shipment quality issues by field:")
    for field_name, count in sorted(issue_counts.items(), key=lambda kv: -kv[1]):
        lines.append(f"  {field_name}: {count}")

    path.write_text("\n".join(lines) + "\n")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--outdir", type=Path, default=DEFAULT_OUTDIR)
    args = parser.parse_args()

    if not args.input.exists():
        raise SystemExit(f"input file not found: {args.input}")

    run(args.input, args.outdir)
    print(f"Wrote normalized CSVs and data_quality_report.csv to {args.outdir}")
    print((args.outdir / "summary.txt").read_text())


if __name__ == "__main__":
    main()
